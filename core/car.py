from datetime import datetime
from decimal import Decimal
from django.db.models import Count, Sum, Case, When, IntegerField
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import Q
from core.models import CargoGroup, Product, ProductCategory
from openpyxl.utils import get_column_letter
from django.db.models import Sum, Count, Q
from openpyxl.cell import MergedCell
from django.contrib.auth.decorators import login_required
@login_required
def export_cargo_report(request, cargo_group_id):
    cargo = get_object_or_404(CargoGroup, id=cargo_group_id)
    
    # Фильтр по статусу товаров
    status_filter = request.GET.get('status', 'all')
    
    # Базовый запрос
    products = Product.objects.filter(cargo_group=cargo).select_related(
        'client', 'category', 'cashbox'
    )
    
    # Применяем фильтр
    if status_filter == 'sold':
        products = products.filter(status='sold')
    elif status_filter == 'unsold':
        products = products.filter(Q(status='kassa') | Q(status='sklad'))
    
    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Рейс {cargo.trip_number}"
    
    # Стили
    header_style = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Заголовок отчета
    ws.merge_cells('A1:Q1')
    ws['A1'] = f"Отчет по рейсу: {cargo.trip_number} ({cargo.date.strftime('%d.%m.%Y')})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Подзаголовок
    ws.merge_cells('A2:Q2')
    status_text = {
        'all': 'Все товары',
        'sold': 'Только проданные',
        'unsold': 'Только не проданные'
    }.get(status_filter, 'Все товары')
    ws['A2'] = f"Статус: {status_text} | Машина: {cargo.vehicle_number}"
    ws['A2'].font = Font(italic=True)
    
    # Заголовки таблицы
    headers = [
        "ID", "Клиент", "Телефон", "Категория",
        "Товар", "Касса", "Кол-во мест", 
        "Вес (кг)", "Кол-во (шт)", "Цена ($)", 
        "Сумма ($)", "Пломба", "Статус", 
        "Дата продажи", "Комментарий"
    ]
    ws.append(headers)
    
    # Форматирование заголовков
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_style
        cell.fill = header_fill
        cell.border = border
    
    # Заполняем данные
    for p in products:
        ws.append([
            p.id,
            p.client.full_name,
            p.client.phone,
            p.category.name if p.category else "",
            p.name,
            p.cashbox.name if p.cashbox else "",
            p.quantity_places,
            p.quantity_kg,
            p.quantity_units,
            p.price,
            p.total_price,
            p.plomb_number,
            p.get_status_display(),
            p.sale_date.strftime("%d.%m.%Y") if p.sale_date else "",
            ""  # Пустое поле для комментариев
        ])
    
    # Форматирование числовых колонок
    for row in ws.iter_rows(min_row=5):
        for col_idx in [7, 8, 9, 10, 11]:  # Колонки с числами
            if row[col_idx-1].value is not None:
                row[col_idx-1].number_format = '#,##0.00'
                row[col_idx-1].border = border
    
    # Итоговая строка
    if products.exists():
        last_row = ws.max_row + 1
        ws.append([
            "Итого:", "", "", "", "", "",
            0,#sum(p.quantity_places or 0  for p in products),
            sum(p.quantity_kg or 0 for p in products),
            sum(p.quantity_units or 0 for p in products),
            "",
            sum(p.total_price for p in products),
            "", "", "", ""
        ])
        
        # Стиль для итогов
        for cell in ws[last_row]:
            cell.font = Font(bold=True)
            cell.border = border
        
        # Средняя цена
        total_kg = sum(p.quantity_kg for p in products)
        if total_kg > 0:
            avg_price = sum(p.total_price for p in products) / Decimal(str(total_kg))
            ws.cell(row=last_row, column=10, value=avg_price).number_format = '#,##0.00'
    
    # Автоширина колонок (исправленная версия)
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # Получаем букву колонки
    
        # Пропускаем объединённые ячейки
        if isinstance(column[0], MergedCell):
            continue
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
    
        # Устанавливаем ширину только для необъединённых колонок
        if max_length > 0:
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

    
    # Формируем ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': (
                f'attachment; filename="cargo_report_{cargo.trip_number}_'
                f'{datetime.now().strftime("%Y%m%d")}.xlsx"'
            )
        },
    )
    wb.save(response)
    return response



@login_required
def cargo_group_detail(request, pk):
    cargo_group = get_object_or_404(CargoGroup.objects.prefetch_related(
        'products__client',
        'products__category'
    ), pk=pk)
    
    # Статистика
    stats = cargo_group.products.aggregate(
        total_weight=Sum('quantity_kg'),
        total_sum=Sum('total_price'),
        sold_count=Count('id', filter=Q(status='sold')),
        kassa_count=Count('id', filter=Q(status='kassa')),
        sklad_count=Count('id', filter=Q(status='sklad'))
    )
    
    # Все категории для фильтра
    categories = ProductCategory.objects.annotate(
        product_count=Count('product', filter=Q(product__cargo_group=cargo_group))
    ).filter(product_count__gt=0)
    
    context = {
        'cargo_group': cargo_group,
        'total_weight': stats['total_weight'],
        'total_sum': stats['total_sum'],
        'sold_count': stats['sold_count'],
        'kassa_count': stats['kassa_count'],
        'sklad_count': stats['sklad_count'],
        'categories': categories
    }
    return render(request, 'cargo_group_detail.html', context)

from django.shortcuts import render
from .models import CargoGroup, Debt

from django.db.models import Sum, Count, Case, When, IntegerField
# views.py
from django.db.models import Sum, Count, Case, When, IntegerField, Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
# views.py
from django.shortcuts import render, get_object_or_404
from django.db.models import Sum, Count, Case, When, IntegerField
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from .models import CargoGroup

@login_required
def cargo_group_list(request):
    cargo_groups = CargoGroup.objects.annotate(
        total_weight=Sum('products__quantity_kg'),
        total_value=Sum('products__total_price'),
        transport_cost_total=Sum('products__transport_cost'),
        sold_count=Count(
            Case(
                When(products__status='sold', then=1),
                output_field=IntegerField()
            )
        )
    ).order_by('-date')
    
    return render(request, 'cargo_group_list.html', {
        'cargo_groups': cargo_groups
    })
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell

@login_required
def export_cargo_groups_report(request):
    if request.method == 'POST':
        selected_groups = request.POST.getlist('selected_groups')
        if not selected_groups:
            return HttpResponse("Не выбрано ни одного рейса", status=400)
            
        groups = CargoGroup.objects.filter(id__in=selected_groups).annotate(
            total_weight=Sum('products__quantity_kg'),
            total_value=Sum('products__total_price'),
            transport_cost_total=Sum('products__transport_cost')
        ).prefetch_related(
            'products__cashbox__category',
            'products__client',
            'products__category'
        )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет по рейсам"
        
        # Заголовок отчета
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = "Отчет по выбранным рейсам"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Шапка таблицы
        headers = [
            "№ Рейса", "Номер машины", "Дата", 
            "Общий вес (кг)", "Стоимость товаров ($)", 
            "Транспортные расходы ($)", "Итого ($)",
            "Кассы (Категория: Название)", "Категории товаров", 
            "Сумма в кассах ($)", "Долги клиентов ($)"
        ]
        ws.append(headers)
        
        # Форматирование заголовков
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
        
        # Данные
        total_sum = 0
        transport_total = 0
        weight_total = 0
        
        for group in groups:
            # Собираем информацию о кассах и категориях
            cashboxes_info = {}
            categories_info = {}
            client_debts = 0
            
            for product in group.products.all():
                # Информация о кассах с указанием категории
                if product.cashbox:
                    cashbox_key = f"{product.cashbox.category.name}: {product.cashbox.name}"
                    cashboxes_info[cashbox_key] = cashboxes_info.get(cashbox_key, 0) + (product.total_price or 0)
                
                # Информация о категориях товаров
                if product.category:
                    category_name = product.category.name
                    categories_info[category_name] = categories_info.get(category_name, 0) + (product.total_price or 0)
                
                # Долги клиентов
                if product.client:
                    client_debts += Debt.objects.filter(
                        client=product.client, 
                        paid=False,
                        is_ours=False
                    ).aggregate(Sum('amount'))['amount__sum'] or 0
            
            group_total = (group.total_value or 0) + Decimal(str((group.transport_cost_total or 0)))
            
            # Формируем строки для касс и категорий
            cashboxes_str = "\n".join([f"{k}: {v}$" for k, v in sorted(cashboxes_info.items())])
            categories_str = "\n".join([f"{k}: {v}$" for k, v in sorted(categories_info.items())])
            
            ws.append([
                group.trip_number,
                group.vehicle_number,
                group.date.strftime("%d.%m.%Y"),
                group.total_weight or 0,
                group.total_value or 0,
                group.transport_cost_total or 0,
                group_total,
                cashboxes_str,
                categories_str,
                sum(cashboxes_info.values()),
                client_debts
            ])
            
            total_sum += group.total_value or 0
            transport_total += group.transport_cost_total or 0
            weight_total += group.total_weight or 0
        
        # Итоговая строка
        ws.append([
            "ИТОГО",
            "",
            "",
            weight_total,
            total_sum,
            transport_total,
            total_sum + Decimal(str(transport_total)),
            "",
            "",
            sum(cashboxes_info.values()) if 'cashboxes_info' in locals() else 0,
            ""
        ])
        
        # Форматирование итогов
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = Font(bold=True)
        
        # Автоподбор ширины колонок с учетом переноса текста
        for col_idx in range(1, len(headers)+1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                try:
                    if cell.value and not isinstance(cell, MergedCell):
                        # Для многострочных ячеек считаем максимальную длину строки
                        if "\n" in str(cell.value):
                            lines = str(cell.value).split("\n")
                            max_line_length = max(len(line) for line in lines)
                            max_length = max(max_length, max_line_length)
                        else:
                            max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Установка переноса текста для ячеек с кассами и категориями
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row-1):
            for cell in [row[7], row[8]]:  # Кассы и категории (колонки H и I)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Формируем ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="cargo_groups_report.xlsx"'}
        )
        wb.save(response)
        return response
    
    return HttpResponse("Метод не разрешен", status=405)