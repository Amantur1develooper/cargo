# from django.shortcuts import render
from openpyxl.utils import get_column_letter
# # Create your views here.
# # views.py
# import openpyxl
# from django.http import HttpResponse
# from .models import Product

# def export_products_in_stock(request):
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Товары на складе"

#     # Заголовки
#     ws.append([
#         "Дата поставки",
#         "Рейс",
#         "Номер машины",
#         "Клиент",
#         "Наименование товара",
#         "Кол-во мест",
#         "Вес (кг)",
#         "Кол-во (шт)",
#         "Цена ($)",
#         "Сумма ($)",
#         "Пломба"
#     ])

#     # Данные
#     products = Product.objects.filter(status='sklad').select_related('client', 'cargo_group')
#     for p in products:
#         ws.append([
#             p.cargo_group.date if p.cargo_group else "",
#             p.cargo_group.trip_number if p.cargo_group else "",
#             p.cargo_group.vehicle_number if p.cargo_group else "",
#             p.client.full_name,
#             p.name,
#             p.quantity_places,
#             p.quantity_kg,
#             p.quantity_units,
#             p.price,
#             p.total_price,
#             p.plomb_number
#         ])

#     # Ответ
#     response = HttpResponse(
#         content=openpyxl.writer.excel.save_virtual_workbook(wb),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename=products_in_stock.xlsx'
#     return response
import json
from django.http import HttpResponse
from .models import CargoGroup, CashTransaction, Client, Product
import openpyxl
from io import BytesIO
from django.core.serializers.json import DjangoJSONEncoder
from django.shortcuts import render
from .models import Product
from .forms import ProductFilterForm
from django.db.models import Q
from django.db.models import Q, Sum
from django.http import HttpResponse
import pandas as pd
from datetime import timedelta
from django.shortcuts import render, redirect
from django.db.models import Q, Sum
from django.contrib import messages
from .models import Product, ProductCategory
from core.forms import ProductFilterForm, AssignCategoryForm
from django.http import JsonResponse
from django.views.decorators.http import require_POST

@require_POST
def clear_products_category(request):
    product_ids = request.POST.getlist('product_ids')
    if not product_ids:
        return JsonResponse({'success': False, 'message': 'Не выбраны товары'})
    
    try:
        updated = Product.objects.filter(
            id__in=product_ids,
            status='sklad'
        ).update(category=None)
        
        return JsonResponse({
            'success': True,
            'message': f'Категория снята с {updated} товаров'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Ошибка: {str(e)}'
        })
        
from django.shortcuts import render, redirect
from django.db.models import Q, Sum, Count
from django.contrib import messages
from .models import Product, ProductCategory
from .forms import ProductFilterForm, AssignCategoryForm


def products_in_stock_view(request):
    form = ProductFilterForm(request.GET or None)
    products = Product.objects.filter(status='sklad').select_related('client', 'cargo_group', 'category')
    
    if form.is_valid():
        client = form.cleaned_data.get('client')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        search_query = request.GET.get('search', '')
        trip_number = request.GET.get('trip_number', '')

        if client:
            products = products.filter(client=client)
        if start_date:
            products = products.filter(cargo_group__date__gte=start_date)
        if end_date:
            products = products.filter(cargo_group__date__lte=end_date)
        if search_query:
            products = products.filter(
                Q(name__icontains=search_query) |
                Q(client__full_name__icontains=search_query) |
                Q(cargo_group__vehicle_number__icontains=search_query)
            )
        if trip_number:
            products = products.filter(cargo_group__trip_number__icontains=trip_number)
    # Обработка формы присвоения категории
    if request.method == 'POST':
        if 'assign_category' in request.POST:
            assign_form = AssignCategoryForm(request.POST)
            print("Form data:", request.POST)  # Отладочный вывод
            print("Selected products:", request.POST.getlist('selected_products'))  # Отладочный вывод
            product_ids = request.POST.getlist('selected_products')  # Это должно работать
    
            if assign_form.is_valid():
                product_ids = request.POST.getlist('selected_products')
                category = assign_form.cleaned_data['category']
                
                print("Received product IDs:", product_ids)  # Для отладки
                if product_ids and category:
                    try:
                        # Преобразуем ID в числа и убираем дубликаты
                        product_ids = list({int(pid) for pid in product_ids if pid.isdigit()})
                        
                        # Обновляем только товары на складе
                        updated = Product.objects.filter(
                            id__in=product_ids,
                            status='sklad'
                        ).update(category=category)
                        
                        messages.success(
                            request, 
                            f"Категория '{category.name}' успешно присвоена {updated} товарам"
                        )
                        return redirect('products_in_stock')
                    
                    except Exception as e:
                        messages.error(request, f"Ошибка при обновлении: {str(e)}")
                else:
                    messages.error(request, "Выберите товары и категорию")
            else:
                messages.error(request, "Пожалуйста, выберите корректную категорию")
        
        elif 'clear_category' in request.POST:
            product_ids = request.POST.getlist('selected_products')
            if product_ids:
                try:
                    product_ids = list({int(pid) for pid in product_ids if pid.isdigit()})
                    updated = Product.objects.filter(
                        id__in=product_ids,
                        status='sklad'
                    ).update(category=None)
                    
                    messages.success(
                        request, 
                        f"Категория снята с {updated} товаров"
                    )
                    return redirect('products_in_stock')
                
                except Exception as e:
                    messages.error(request, f"Ошибка при снятии категории: {str(e)}")
            else:
                messages.error(request, "Выберите товары для снятия категории")
    else:
        assign_form = AssignCategoryForm()

    # Агрегированные данные
    total_stats = products.aggregate(
        total_sum=Sum('total_price'),
        total_weight=Sum('quantity_kg'),
        total_units=Sum('quantity_units'),
        total_items=Count('id')
    )
     # Получаем список уникальных рейсов для фильтра
    trip_numbers = CargoGroup.objects.filter(
        products__status='sklad'
    ).distinct().values_list('trip_number', flat=True)

    context = {
        'products': products,
        'form': form,
        'assign_form': assign_form,
        'total_sum': total_stats['total_sum'] or 0,
        'total_weight': total_stats['total_weight'] or 0,
        'total_units': total_stats['total_units'] or 0,
        'total_items': total_stats['total_items'] or 0,
        'trip_numbers': trip_numbers,
    }
    return render(request, 'products_in_stock.html', context)

def export_products_in_stock(request):
    products = Product.objects.filter(status='sklad').select_related('client', 'cargo_group')
    
    data = []
    for p in products:
        data.append({
            'Дата поставки': p.cargo_group.date.strftime('%d.%m.%Y') if p.cargo_group.date else '',
            'Рейс': p.cargo_group.trip_number or '',
            'Номер машины': p.cargo_group.vehicle_number or '',
            'Клиент': p.client.full_name,
            'Товар': p.name,
            'Категория': p.category.name if p.category else '',
            'Кол-во мест': p.quantity_places,
            'Вес (кг)': p.quantity_kg,
            'Кол-во (шт)': p.quantity_units,
            'Цена ($)': p.price,
            'Сумма ($)': p.total_price,
            'Пломба': p.plomb_number or ''
        })

    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=products_in_stock.xlsx'
    df.to_excel(response, index=False)
    return response
# def products_in_stock_view(request):
#     form = ProductFilterForm(request.GET or None)
#     products = Product.objects.filter(status='sklad').select_related('client', 'cargo_group')

#     if form.is_valid():
#         client = form.cleaned_data.get('client')
#         start_date = form.cleaned_data.get('start_date')
#         end_date = form.cleaned_data.get('end_date')

#         if client:
#             products = products.filter(client=client)
#         if start_date:
#             products = products.filter(cargo_group__date__gte=start_date)
#         if end_date:
#             products = products.filter(cargo_group__date__lte=end_date)

#     context = {
#         'products': products,
#         'form': form,
#     }
#     return render(request, 'products_in_stock.html', context)

# # def products_in_stock_view(request):
# #     products = Product.objects.filter(status='sklad').select_related('client', 'cargo_group')
# #     return render(request, 'products_in_stock.html', {'products': products})
# from django.http import HttpResponse
# import openpyxl
# from io import BytesIO
# from .models import Product

# def export_products_in_stock(request):
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Товары на складе"

#     ws.append([
#         "Дата поставки", "Рейс", "Номер машины", "Клиент",
#         "Наименование товара", "Кол-во мест", "Вес (кг)",
#         "Кол-во (шт)", "Цена ($)", "Сумма ($)", "Пломба"
#     ])

#     products = Product.objects.filter(status='sklad').select_related('client', 'cargo_group')
#     for p in products:
#         ws.append([
#             p.cargo_group.date if p.cargo_group else "",
#             p.cargo_group.trip_number if p.cargo_group else "",
#             p.cargo_group.vehicle_number if p.cargo_group else "",
#             p.client.full_name if p.client else "",
#             p.name,
#             p.quantity_places,
#             p.quantity_kg,
#             p.quantity_units,
#             p.price,
#             p.total_price,
#             p.plomb_number
#         ])

#     buffer = BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)

#     response = HttpResponse(
#         buffer.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename=products_in_stock.xlsx'
#     return response

# def export_products_in_stock(request):
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Товары на складе"

#     # Заголовки
#     ws.append([
#         "Дата поставки",
#         "Рейс",
#         "Номер машины",
#         "Клиент",
#         "Наименование товара",
#         "Кол-во мест",
#         "Вес (кг)",
#         "Кол-во (шт)",
#         "Цена ($)",
#         "Сумма ($)",
#         "Пломба"
#     ])

#     # Данные
#     products = Product.objects.filter(status='sklad').select_related('client', 'cargo_group')
#     for p in products:
#         ws.append([
#             p.cargo_group.date if p.cargo_group else "",
#             p.cargo_group.trip_number if p.cargo_group else "",
#             p.cargo_group.vehicle_number if p.cargo_group else "",
#             p.client.full_name,
#             p.name,
#             p.quantity_places,
#             p.quantity_kg,
#             p.quantity_units,
#             p.price,
#             p.total_price,
#             p.plomb_number
#         ])

#     # Сохраняем в память
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename=products_in_stock.xlsx'

#     buffer = BytesIO()
#     wb.save(buffer)
#     response.write(buffer.getvalue())
#     return response



# views.py

from .models import Currency

def generate_invoice(request):
    ids = request.GET.get('ids')
    id_list = ids.split(',') if ids else []
    products = Product.objects.filter(id__in=id_list)

    total_usd = sum(p.total_price for p in products if p.total_price)

    # Получаем последний курс из базы
    # currency = Currency.objects.last()
    # som_rate = currency.usd_to_som if currency else 88
    total_som = total_usd 

    return render(request, "invoice.html", {
        "products": products,
        "total_usd": total_usd,
        "total_som": total_som,
        "som_rate":total_som ,
    })

# def generate_invoice(request):
#     ids = request.GET.get('ids')
#     id_list = ids.split(',') if ids else []
#     products = Product.objects.filter(id__in=id_list)
#     total_usd = sum(p.total_price for p in products if p.total_price)
#     som_rate = 88  # пример курса
#     total_som = total_usd * som_rate
#     return render(request, "invoice.html", {
#         "products": products,
#         "total_usd": total_usd,
#         "total_som": total_som,
#     })
# def generate_invoice(request):
#     ids = request.GET.get('ids', '')
#     id_list = ids.split(',') if ids else []
#     products = Product.objects.filter(id__in=id_list)

#     if not products.exists():
#         return render(request, 'invoice/empty.html')

#     return render(request, 'invoice/invoice.html', {
#         'products': products,
#     })

from django.shortcuts import render
from .models import Product, CashBox, Debt, ProductCategory
from django.db.models.functions import TruncDay, TruncMonth
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from datetime import timedelta
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import timedelta
from django.contrib import messages
from django.db import transaction
from .forms import WithdrawalForm
from django.db.models.functions import Coalesce
from django.db.models import Sum, Q, Count
from django.db.models.functions import Coalesce, TruncDay
from django.db.models import DecimalField
from decimal import Decimal
def dashboard(request):
    # 1. Основные финансовые показатели (включая изъятия)
    cash_data = CashTransaction.objects.aggregate(
        income=Coalesce(Sum('amount', filter=Q(type='income'), output_field=DecimalField()), Decimal(0)),
        expense=Coalesce(Sum('amount', filter=Q(type='expense'), output_field=DecimalField()), Decimal(0)),
        withdrawals=Coalesce(Sum('amount', filter=Q(type='withdrawal'), output_field=DecimalField()), Decimal(0))
    )
    
    # Рассчитываем баланс с учетом изъятий
    cash_balance = cash_data['income'] - cash_data['expense'] - cash_data['withdrawals']

    # 2. Товары и категории
    products_data = Product.objects.aggregate(
        total_value=Coalesce(Sum('total_price', filter=Q(status='sklad'), output_field=DecimalField()), Decimal(0)),
        tmp_value=Coalesce(Sum('total_price', filter=Q(category__name='ТМП', status='sklad'), output_field=DecimalField()), Decimal(0)),
        hoz_value=Coalesce(Sum('total_price', filter=Q(category__name='Хозтовары', status='sklad'), output_field=DecimalField()), Decimal(0))
    )

    # 3. Долги
    debts = Debt.objects.aggregate(
        to_us=Coalesce(Sum('amount', filter=Q(paid=False, is_ours=False), output_field=DecimalField()), Decimal(0)),
        ours=Coalesce(Sum('amount', filter=Q(paid=False, is_ours=True), output_field=DecimalField()), Decimal(0))
    )

    # 4. Данные для графиков (последние 30 дней)
    date_range = timezone.now() - timedelta(days=30)
    
    # График продаж (исключаем изъятия)
    sales = (
        CashTransaction.objects
        .filter(type='income', created_at__gte=date_range)
        .annotate(date=TruncDay('created_at'))
        .values('date')
        .annotate(total=Sum('amount', output_field=DecimalField()))
        .order_by('date')
    )

    # Топ клиентов
    top_clients = (
        Client.objects
        .annotate(total_products=Count('products'))
        .order_by('-total_products')[:5]
    )

    # Рассчитываем все значения
    cash_balance = float(cash_balance)
    total_value = float(products_data['total_value'])
    debt_to_us = float(debts['to_us'])
    our_debt = float(debts['ours'])

    context = {
        # Финансы (теперь с учетом изъятий)
        'cash_balance': round(cash_balance, 2),
        'cash_income': round(float(cash_data['income']), 2),
        'cash_expense': round(float(cash_data['expense']), 2),
        'cash_withdrawals': round(float(cash_data['withdrawals']), 2),
        
        # Товары
        'product_total': round(total_value, 2),
        'tmp_total': round(float(products_data['tmp_value']), 2),
        'hoz_total': round(float(products_data['hoz_value']), 2),
        
        # Долги
        'debt_to_us': round(debt_to_us, 2),
        'debt_ours': round(our_debt, 2),
        'net_assets': round(cash_balance + total_value + debt_to_us - our_debt, 2),
        
        # Графики
        'sales_data': list(sales),
        'top_clients': top_clients,
    }
    # График продаж за последние 30 дней
    date_range = timezone.now() - timedelta(days=30)
    sales_data = (
        CashTransaction.objects
        .filter(type='income', created_at__gte=date_range)
        .annotate(day=TruncDay('created_at'))
        .values('day')
        .annotate(total=Sum('amount', output_field=DecimalField()))
        .order_by('day')
    )
    
    # Заполняем пропущенные дни нулями
    full_sales_data = []
    current_date = date_range.date()
    today = timezone.now().date()
    
    while current_date <= today:
        day_data = next((item for item in sales_data if item['day'].date() == current_date), None)
        if day_data:
            full_sales_data.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'total': float(day_data['total'])
            })
        else:
            full_sales_data.append({
                'date': current_date.strftime('%Y-%m-%d'),
                'total': 0.0
            })
        current_date += timedelta(days=1)
    
    context['sales_data_json'] = json.dumps(full_sales_data, cls=DjangoJSONEncoder)
    
    # Обработка формы изъятия средств
    withdrawal_form = WithdrawalForm()
    
    if request.method == 'POST' and 'withdraw' in request.POST:
        withdrawal_form = WithdrawalForm(request.POST)
        if withdrawal_form.is_valid():
            try:
                with transaction.atomic():
                    amount = withdrawal_form.cleaned_data['amount']
                    description = withdrawal_form.cleaned_data['description'] or "Изъятие средств администратором"
                    
                    # Проверка достаточности средств (учитываем текущий баланс)
                    if amount > cash_balance:
                        messages.error(request, "Недостаточно средств на счету")
                    else:
                        # Создаем запись о выводе средств
                        CashTransaction.objects.create(
                            amount=amount,
                            type=CashTransaction.WITHDRAWAL,
                            description=description,
                            created_by=request.user
                        )
                        messages.success(request, f"Успешно изъято {amount} из кассы")
                        return redirect('dashboard')
            
            except Exception as e:
                messages.error(request, f"Ошибка при изъятии средств: {str(e)}")
    
    context.update({
        'withdrawal_form': withdrawal_form,
    })
    
    return render(request, 'dashboard/new_dashboard.html', context)
def get_context_data(self, **kwargs):
    context = super().get_context_data(**kwargs)
    context['sold_products'] = Product.objects.filter(status='sold').count()
    return context
from django.shortcuts import render
from .models import Product, ProductCategory

from django.shortcuts import render, get_object_or_404
from django.db.models import Sum, Count, Q
from .models import ProductCategory, Product
from django.core.paginator import Paginator
from django.shortcuts import render, get_object_or_404
from django.db.models import Sum, Count, Q
from django.core.paginator import Paginator
from .models import ProductCategory, Product

def products_by_category(request, category_name):
    # Получаем категорию (без prefetch_related, так как это ForeignKey в Product)
    cashboxes = CashBox.objects.filter(category__name=category_name)
    category = get_object_or_404(
        ProductCategory,
        name__iexact=category_name
    )
    
    # Базовый запрос с оптимизацией
    products = Product.objects.filter(category=category)\
        .select_related('client', 'cargo_group', 'category')\
    
    # Фильтрация по статусу
    status_filter = request.GET.get('status')
    if status_filter in ['sklad', 'way', 'kassa', 'sold']:
        products = products.filter(status=status_filter)
    
    # Поиск
    search_query = request.GET.get('q')
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(client__full_name__icontains=search_query)
        )
    
    # Статистика
    category_stats = {
        'total_products': products.count(),
        'total_weight': products.aggregate(Sum('quantity_kg'))['quantity_kg__sum'] or 0,
        'total_value': products.aggregate(Sum('total_price'))['total_price__sum'] or 0,
        'sold_count': products.filter(status='sold').count(),
    }
    
    # Пагинация
    page_number = request.GET.get('page', 1)
    paginator = Paginator(products, 25)
    page_obj = paginator.get_page(page_number)
    
   
    context = {
        'category': category,
        'products': page_obj,
        'category_stats': category_stats,
        'current_status': status_filter,
        'search_query': search_query,

        'cashboxes': cashboxes,
    }
    
    return render(request, 'products/category_products.html', context)
    
from django.views.decorators.http import require_POST
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
def category_cashboxes(request, category_name):
    category = get_object_or_404(ProductCategory, name__iexact=category_name)
    cashboxes = CashBox.objects.filter(category=category)
    products = Product.objects.filter(category=category, status='sklad')
    
    return render(request, 'cashboxes/category_cashboxes.html', {
        'category': category,
        'cashboxes': cashboxes,
        'products': products,
        'ProductCategory':ProductCategory.objects.all()
    })

@require_POST
def move_products_to_cashbox(request):
    cashbox_id = request.POST.get('cashbox')
    product_ids = request.POST.getlist('product_ids')
    cashbox = get_object_or_404(CashBox, id=cashbox_id)
    
    Product.objects.filter(id__in=product_ids).update(
        cashbox=cashbox,
        status='kassa'
    )
    
    messages.success(request, f'Товары перемещены в кассу "{cashbox.name}"')
    return redirect('category_cashboxes', category_name=cashbox.category.name)

from django.utils import timezone

# views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Sum
from .models import CashBox, Product, Currency, Debt
from django.contrib import messages
# views.pyfrom django.utils import timezone
from decimal import Decimal

def cashbox_detail(request, cashbox_id):
    cashbox = get_object_or_404(CashBox, id=cashbox_id)
    products = Product.objects.filter(cashbox=cashbox, status='kassa')
    current_currency = Currency.objects.last()
    
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        paid_amount_som = Decimal(request.POST.get('paid_amount', 0))
        
        product = get_object_or_404(Product, id=product_id)
        paid_amount_usd = paid_amount_som 
        
        # Проверяем, есть ли существующий долг по этому товару
        existing_debt = Debt.objects.filter(product=product, paid=False).first()
        
        if existing_debt:
            # Если долг уже есть, обновляем его
            remaining_debt = existing_debt.amount - paid_amount_usd
            if remaining_debt <= 0:
                existing_debt.paid = True
                existing_debt.save()
                # Используем функцию sell_product для завершения продажи
                return sell_product(request, product_id)
            else:
                existing_debt.amount = remaining_debt
                existing_debt.save()
                messages.warning(request, 
                    f"Частичная оплата. Остаток долга: ${remaining_debt:.2f}")
        else:
            if paid_amount_usd < product.total_price:
                # Создаем запись о долге
                paid_amount_usd = Decimal(str(paid_amount_usd))
                debt_amount = product.total_price - paid_amount_usd 
                paid_amount_usd = float(str(paid_amount_usd))
                Debt.objects.create(
                    client=product.client,
                    amount=debt_amount,
                    is_ours=False,
                    product=product
                )
                messages.warning(request, 
                    f"Частичная оплата. Остаток долга: ${debt_amount:.2f}")
            else:
                # Используем функцию sell_product для завершения продажи
                return sell_product(request, product_id)
        
        return redirect('cashbox_detail', cashbox_id=cashbox_id)
    
    # Добавляем информацию о долгах к товарам
    products_with_debt = []
    for product in products:
        debt = Debt.objects.filter(product=product, paid=False).first()
        products_with_debt.append({
            'product': product,
            'debt': debt
        })
    
    return render(request, 'cashboxes/cashbox_detail.html', {
        'cashbox': cashbox,
        'products_with_debt': products_with_debt,
        'current_currency': current_currency,
    })

@require_POST
def sell_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    
    # Обновляем статус и дату продажи
    product.status = 'sold'
    product.sale_date = timezone.now()
    product.save()
    
    # Создаем запись о транзакции
    CashTransaction.objects.create(
        cashbox=product.cashbox,
        type='income',
        amount=product.total_price,
        description=f"Продажа товара: {product.name}",
        client=product.client,
        product=product,
        created_by=request.user
    )
    
    messages.success(request, f'Товар "{product.name}" успешно продан!')
    return redirect('cashbox_detail', cashbox_id=product.cashbox.id)

@require_POST
def return_to_warehouse(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    cashbox_id = product.cashbox.id if product.cashbox else None
    product.status = 'sklad'
    product.cashbox = None
    product.save()
    
    if cashbox_id:
        messages.success(request, f'Товар "{product.name}" возвращён на склад.')
        return redirect('cashbox_detail', cashbox_id=cashbox_id)
    else:
        messages.success(request, f'Товар "{product.name}" возвращён на склад.')
        return redirect('category_cashboxes', category_name=product.category.name)
# views.py
from django.db.models import Sum, Avg, Q
from django.contrib import messages
from datetime import datetime
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
def sales_report(request):
    # Начальный QuerySet
    sales = CashTransaction.objects.filter(type='income').order_by('-created_at')
    
    # Обработка фильтров
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    search_query = request.GET.get('search', '').strip()
    
    # Фильтрация по дате
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            sales = sales.filter(created_at__date__gte=start_date)
        except ValueError:
            messages.error(request, "Неверный формат начальной даты")
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            sales = sales.filter(created_at__date__lte=end_date)
        except ValueError:
            messages.error(request, "Неверный формат конечной даты")
    
    # Фильтрация по поисковому запросу
    if search_query:
        sales = sales.filter(
            Q(client__full_name__icontains=search_query) |
            Q(product__name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(cashbox__name__icontains=search_query)
        )
    
    # Расчет агрегированных данных
    total_amount = sales.aggregate(total=Sum('amount'))['total'] or 0
    avg_amount = sales.aggregate(avg=Avg('amount'))['avg'] or 0
    
    # Пагинация
    paginator = Paginator(sales, 25)  # 25 записей на страницу
    page_number = request.GET.get('page')
    try:
        sales = paginator.page(page_number)
    except PageNotAnInteger:
        sales = paginator.page(1)
    except EmptyPage:
        sales = paginator.page(paginator.num_pages)
    
    context = {
        'sales': sales,
        'total_amount': total_amount,
        'avg_amount': avg_amount,
    }
    
    return render(request, 'reports/sales.html', context)


# views.py
from django.db.models import Count, Sum

def client_list(request):
    clients = Client.objects.annotate(
        product_count=Count('products'),
        unpaid_products=Count('products', filter=Q(products__status='kassa')),
        total_debt=Sum('debts__amount', filter=Q(debts__paid=False))
    ).order_by('-total_debt')
    
    return render(request, 'clients/client_list.html', {'clients': clients})

def client_detail(request, client_id):
    client = get_object_or_404(Client, id=client_id)
    products = client.products.all()
    debts = client.debts.filter(paid=False)
    
    return render(request, 'clients/client_detail.html', {
        'client': client,
        'products': products,
        'debts': debts,
        'stats': client.get_product_stats()
    })
    
    
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

def export_categories_report(request):
    # Создаем Excel-файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по категориям"

    # Заголовки
    headers = [
        "Категория", 
        "Кол-во товаров", 
        "Общий вес (кг)", 
        "Общая сумма ($)",
        "На складе",
        "На кассе",
        "Продано"
    ]
    
    ws.append(headers)

    # Стили для заголовков
    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

    # Получаем данные
    from django.db.models import Count, Sum, Case, When, IntegerField
    categories = ProductCategory.objects.annotate(
        total_products=Count('product'),
        total_weight=Sum('product__quantity_kg'),
        total_sum=Sum('product__total_price'),
        on_stock=Count('product', filter=Q(product__status='sklad')),
        on_cashbox=Count('product', filter=Q(product__status='kassa')),
        sold=Count('product', filter=Q(product__status='sold'))
    )

    # Заполняем данные
    for category in categories:
        ws.append([
            category.name,
            category.total_products,
            category.total_weight or 0,
            category.total_sum or 0,
            category.on_stock,
            category.on_cashbox,
            category.sold
        ])

    # Авто-ширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Сохраняем файл
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=categories_report_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    wb.save(response)
    return response
from openpyxl.styles import PatternFill
def export_cashboxes_report(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по кассам"

    # Заголовки
    headers = [
        "Касса", 
        "Категория",
        "Кол-во товаров", 
        "Общая сумма ($)",
        "Средняя цена ($)"
    ]
    ws.append(headers)

    # Стили
    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')

    # Данные
    from django.db.models import Avg
    cashboxes = CashBox.objects.select_related('category').annotate(
        products_count=Count('product'),
        total_sum=Sum('product__total_price'),
        avg_price=Avg('product__price')
    )

    for cashbox in cashboxes:
        ws.append([
            cashbox.name,
            cashbox.category.name,
            cashbox.products_count,
            cashbox.total_sum or 0,
            cashbox.avg_price or 0
        ])

    # Форматирование чисел
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = '#,##0.00'

    # Авто-ширина
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=cashboxes_report_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    wb.save(response)
    return response

from django.shortcuts import get_object_or_404
from django.db.models import Q
from openpyxl.styles import NamedStyle, Font, Border, Side
# def export_cashbox_detail(request, cashbox_id):
#     cashbox = get_object_or_404(CashBox, id=cashbox_id)
    
#     # Фильтрация
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')
#     status_filter = request.GET.get('status', 'all')
    
#     products = Product.objects.filter(cashbox=cashbox).select_related(
#         'client', 'cargo_group', 'category'
#     )
    
#     if start_date:
#         products = products.filter(cargo_group__date__gte=start_date)
#     if end_date:
#         products = products.filter(cargo_group__date__lte=end_date)
    
#     if status_filter == 'sold':
#         products = products.filter(status='sold')
#     elif status_filter == 'unsold':
#         products = products.filter(status='kassa')

#     # Создаем Excel
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Отчет по кассе"
    
#     # Стили
#     title_style = NamedStyle(name="title")
#     title_style.font = Font(bold=True, size=14)
#     title_style.alignment = Alignment(horizontal='center')
    
#     header_style = NamedStyle(name="header")
#     header_style.font = Font(bold=True, color="FFFFFF")
#     header_style.fill = PatternFill(
#         start_color="4F81BD", 
#         end_color="4F81BD", 
#         fill_type="solid"
#     )
#     header_style.border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )

#     # Основной заголовок
#     ws.merge_cells('A1:Q1')
#     title_cell = ws['A1']
#     title_cell.value = f"{cashbox.category.name} - Касса: {cashbox.name}"
#     title_cell.style = title_style

#     # Подзаголовок с фильтрами
#     ws.merge_cells('A2:Q2')
#     filter_text_parts = []
#     if start_date or end_date:
#         date_range = []
#         if start_date:
#             date_range.append(f"с {start_date}")
#         if end_date:
#             date_range.append(f"по {end_date}")
#         filter_text_parts.append(f"Период: {' '.join(date_range)}")
    
#     if status_filter != 'all':
#         status_text = {
#             'sold': 'Только проданные',
#             'unsold': 'Только на кассе'
#         }.get(status_filter, 'Все')
#         filter_text_parts.append(f"Статус: {status_text}")
    
#     filter_text = " | ".join(filter_text_parts) if filter_text_parts else "Полный отчет"
#     ws['A2'] = filter_text
#     ws['A2'].font = Font(italic=True)

#     # Пустая строка перед таблицей
#     ws.append([])

#     # Заголовки таблицы
#     headers = [
#         "ID", "Клиент", "Телефон", "Город",
#         "Товар", "Дата поставки", "Рейс",
#         "Номер машины", "Кол-во мест", "Вес (кг)", "Кол-во (шт)",
#         "Цена ($)", "Сумма ($)", "Пломба", "Статус",
#         "Дата продажи"
#     ]
#     ws.append(headers)
    
#     # Применяем стили к заголовкам
#     for cell in ws[4]:  # Заголовки в 4 строке
#         cell.style = header_style
    
#     # Заполнение данных
#     for p in products:
#         ws.append([
#             p.id,
#             p.client.full_name,
#             p.client.phone,
#             p.client.city,
#             p.name,
#             p.cargo_group.date.strftime("%d.%m.%Y") if p.cargo_group and p.cargo_group.date else "",
#             p.cargo_group.trip_number if p.cargo_group else "",
#             p.cargo_group.vehicle_number if p.cargo_group else "",
#             p.quantity_places,
#             p.quantity_kg,
#             p.quantity_units,
#             p.price,
#             p.total_price,
#             p.plomb_number,
#             p.get_status_display(),
#             p.sale_date.strftime("%d.%m.%Y") if p.sale_date else ""
#         ])

#     # Границы для всех ячеек с данными
#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
    
#     # Форматирование числовых колонок
#     number_columns = ['K', 'L', 'M', 'N']  # Вес, Кол-во, Цена, Сумма
#     for col in number_columns:
#         for cell in ws[col]:
#             if cell.row > 4:  # Пропускаем заголовки
#                 cell.number_format = '#,##0.00'
#                 cell.border = border

#     # Итоговая строка
#     if products.exists():
#         last_row = ws.max_row + 1
#         ws.append([
#             "Итого:", "", "", "", "", "", "", "",
#             sum(p.quantity_places for p in products),
#             sum(p.quantity_kg for p in products),
#             sum(p.quantity_units for p in products),
#             "",
#             sum(p.total_price for p in products),
#             "", "", ""
#         ])
        
#         # Стиль для итоговой строки
#         for cell in ws[last_row]:
#             cell.font = Font(bold=True)
#             cell.border = border
        
#         # Средняя цена (отдельный расчет)
#         avg_price = sum(p.total_price for p in products) / sum(p.quantity_kg for p in products) if sum(p.quantity_kg for p in products) > 0 else 0
#         ws.cell(row=last_row, column=12, value=avg_price).number_format = '#,##0.00'

#     # Автоматическая ширина колонок
#     for column in ws.columns:
#         max_length = 0
#         column_letter = column[0].column_letter
#         for cell in column:
#             try:
#                 value = str(cell.value) if cell.value is not None else ""
#                 # Учитываем все строки, кроме заголовков отчета
#                 if cell.row > 2:  
#                     if len(value) > max_length:
#                         max_length = len(value)
#             except:
#                 pass
#         adjusted_width = (max_length + 2) * 1.1
#         ws.column_dimensions[column_letter].width = adjusted_width

#     # Условное форматирование для статусов
#     status_col = 15  # Колонка статуса (O)
#     for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
#         status_cell = row[status_col - 1]  # -1 потому что индексация с 0
#         if status_cell.value == "Продано":
#             status_cell.font = Font(color="00AA00")  # Зеленый
#         elif status_cell.value == "На кассе":
#             status_cell.font = Font(color="0000FF")  # Синий
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.db.models import Sum

def export_cashbox_detail(request, cashbox_id):
    cashbox = get_object_or_404(CashBox, id=cashbox_id)
    
    # Фильтрация
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status_filter = request.GET.get('status', 'all')
    
    products = Product.objects.filter(cashbox=cashbox).select_related(
        'client', 'cargo_group', 'category'
    )
    
    if start_date:
        products = products.filter(cargo_group__date__gte=start_date)
    if end_date:
        products = products.filter(cargo_group__date__lte=end_date)
    
    if status_filter == 'sold':
        products = products.filter(status='sold')
    elif status_filter == 'unsold':
        products = products.filter(status='kassa')

    # Создаем Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет по кассе"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Заголовок отчета
    ws.merge_cells('A1:Q1')
    ws['A1'] = f"{cashbox.category.name} - Касса: {cashbox.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Подзаголовок с фильтрами
    ws.merge_cells('A2:Q2')
    filter_text = []
    if start_date or end_date:
        filter_text.append(f"Период: {start_date or ''} - {end_date or ''}")
    if status_filter != 'all':
        filter_text.append(f"Статус: {'Проданные' if status_filter == 'sold' else 'На кассе'}")
    
    ws['A2'] = " | ".join(filter_text) if filter_text else "Полный отчет"
    ws['A2'].font = Font(italic=True)

    # Заголовки таблицы
    headers = [
        "ID", "Клиент", "Телефон", "Город", "Товар", 
        "Дата поставки", "Рейс", "Номер машины", 
        "Кол-во мест", "Вес (кг)", "Кол-во (шт)", 
        "Цена ($)", "Сумма ($)", "Пломба", "Статус", "Дата продажи"
    ]
    ws.append(headers)
    
    # Стили для заголовков
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Заполнение данных
    for p in products:
        row = [
            p.id,
            p.client.full_name,
            p.client.phone,
            p.client.city,
            p.name,
            p.cargo_group.date.strftime("%d.%m.%Y") if p.cargo_group and p.cargo_group.date else "",
            p.cargo_group.trip_number if p.cargo_group else "",
            p.cargo_group.vehicle_number if p.cargo_group else "",
            p.quantity_places,
            p.quantity_kg,
            p.quantity_units,
            p.price,
            p.total_price,
            p.plomb_number,
            p.get_status_display(),
            p.sale_date.strftime("%d.%m.%Y") if p.sale_date else ""
        ]
        ws.append(row)

    # Форматирование числовых колонок
    for row in ws.iter_rows(min_row=5):
        for col_idx in [9, 10, 11, 12, 13]:  # Колонки с числами
            if row[col_idx-1].value is not None:
                row[col_idx-1].number_format = '#,##0.00'
                row[col_idx-1].border = border

    # Итоговая строка с правильными расчетами
    if products.exists():
        total_row = [
            "Итого:", "", "", "", "", "", "", "",
            sum(p.quantity_places or 0 for p in products),
            sum(p.quantity_kg or 0 for p in products),
            sum(p.quantity_units or 0 for p in products),
            "",
            sum(p.total_price for p in products),
            "", "", ""
        ]
        ws.append(total_row)
        
        # Форматирование итоговой строки
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.border = border

        # Средняя цена
        total_kg = sum(Decimal(str(p.quantity_kg)) if p.quantity_kg else Decimal('0') for p in products)
        if total_kg > 0:
            total_price = sum(Decimal(str(p.total_price)) if p.total_price else Decimal('0') for p in products)
            avg_price = float(total_price / total_kg)  # Преобразуем в float после деления Decimal
            ws.cell(row=ws.max_row + 1, column=1, value="Средняя цена за кг:")
            ws.cell(row=ws.max_row, column=2, value=round(avg_price, 2))
    
        # total_kg = sum(p.quantity_kg for p in products)
        # if total_kg > 0:
        #     avg_price = sum(p.total_price for p in products) / total_kg
        #     ws.cell(row=ws.max_row, column=12, value=avg_price).number_format = '#,##0.00'

    # Автоматическая ширина колонок
    for col_idx in range(1, len(headers)+1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Пропускаем объединенные ячейки
        is_merged = any(
            column_letter in merged_range.coord 
            for merged_range in ws.merged_cells.ranges
        )
        
        if not is_merged:
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
                cell = row[col_idx-1]
                try:
                    value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(value))
                except:
                    pass
                    
            ws.column_dimensions[column_letter].width = (max_length + 2) * 1.1

    # Настройка HTTP-ответа
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': (
                f'attachment; '
                f'filename="{cashbox.category.name}_{cashbox.name}_report_'
                f'{datetime.now().strftime("%Y-%m-%d")}.xlsx"'
            )
        },
    )
    wb.save(response)
    return response
#     # Сохраняем файл
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = (
#         f'attachment; filename={cashbox.category.name}_{cashbox.name}_report_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
#     )
#     wb.save(response)
#     return response
# def export_cashbox_detail(request, cashbox_id):
#     cashbox = get_object_or_404(CashBox, id=cashbox_id)
    
#     # Получаем параметры фильтрации
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')
#     status_filter = request.GET.get('status', 'all')  # all, sold, unsold
    
#     # Базовый запрос
#     products = Product.objects.filter(cashbox=cashbox).select_related(
#         'client', 'cargo_group', 'category'
#     )
    
#     # Применяем фильтры
#     if start_date:
#         products = products.filter(cargo_group__date__gte=start_date)
#     if end_date:
#         products = products.filter(cargo_group__date__lte=end_date)
    
#     if status_filter == 'sold':
#         products = products.filter(status='sold')
#     elif status_filter == 'unsold':
#         products = products.filter(status='kassa')
    
#     # Создаем Excel-файл
#     wb = Workbook()
#     ws = wb.active
#     ws.title = f"Касса {cashbox.name}"
    
#     # Стили для Excel
#     header_style = NamedStyle(name="header")
#     header_style.font = Font(bold=True, color="FFFFFF")
#     header_style.fill = openpyxl.styles.PatternFill(
#         start_color="4F81BD", 
#         end_color="4F81BD", 
#         fill_type="solid"
#     )
#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
    
#     # Заголовки
#     headers = [
#         "ID", "Клиент", "Телефон", "Город",
#         "Товар", "Категория", "Дата поставки", "Рейс",
#         "Номер машины", "Кол-во мест", "Вес (кг)", "Кол-во (шт)",
#         "Цена ($)", "Сумма ($)", "Пломба", "Статус",
#         "Дата продажи"
#     ]
#     ws.append(headers)
    
#     # Применяем стили к заголовкам
#     for cell in ws[1]:
#         cell.style = header_style
#         cell.border = border
    
#     # Заполняем данные
#     for p in products:
#         ws.append([
#             p.id,
#             p.client.full_name,
#             p.client.phone,
#             p.client.city,
#             p.name,
#             p.category.name if p.category else "",
#             p.cargo_group.date.strftime("%d.%m.%Y") if p.cargo_group and p.cargo_group.date else "",
#             p.cargo_group.trip_number if p.cargo_group else "",
#             p.cargo_group.vehicle_number if p.cargo_group else "",
#             p.quantity_places,
#             p.quantity_kg,
#             p.quantity_units,
#             p.price,
#             p.total_price,
#             p.plomb_number,
#             p.get_status_display(),
#             p.sale_date.strftime("%d.%m.%Y") if p.sale_date else ""
#         ])
   
   

def product_detail(request, pk):
    product = get_object_or_404(Product.objects.select_related(
        'client', 'category', 'cashbox', 'cargo_group'
    ), pk=pk)
    return render(request, 'products/product_detail.html', {'product': product})